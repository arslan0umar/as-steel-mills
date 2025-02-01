from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.textinput import TextInput
from kivy.uix.spinner import Spinner
from kivy.uix.popup import Popup
from kivy.uix.scrollview import ScrollView
from kivy.uix.gridlayout import GridLayout
from kivy.uix.checkbox import CheckBox
from kivy.uix.treeview import TreeView, TreeViewLabel
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook

class SteelInventoryApp(App):
    def build(self):
        self.title = "AS Steel Mills"
        self.parties = []
        self.load_parties()
        return self.main_screen()

    def load_parties(self):
        if os.path.exists("parties.txt"):
            with open("parties.txt", "r") as file:
                self.parties = file.read().splitlines()

    def save_parties(self):
        with open("parties.txt", "w") as file:
            file.write("\n".join(self.parties))

    def main_screen(self):
        layout = BoxLayout(orientation="vertical", padding=10, spacing=10)

        title = Label(text="AS Steel Mills", font_size=24, bold=True)
        layout.add_widget(title)

        enter_data_button = Button(text="Enter New Data", size_hint=(1, 0.2))
        enter_data_button.bind(on_press=self.data_entry_screen)
        layout.add_widget(enter_data_button)

        open_ledger_button = Button(text="Open Ledger", size_hint=(1, 0.2))
        open_ledger_button.bind(on_press=self.open_ledger_screen)
        layout.add_widget(open_ledger_button)

        return layout

    def data_entry_screen(self, instance):
        self.clear_screen()

        layout = BoxLayout(orientation="vertical", padding=10, spacing=10)

        title = Label(text="AS Steel Inventory", font_size=20, bold=True)
        layout.add_widget(title)

        # Party selection
        self.party_spinner = Spinner(text="Select Party", values=self.parties, size_hint=(1, 0.1))
        layout.add_widget(self.party_spinner)

        add_party_button = Button(text="Add Party", size_hint=(1, 0.1))
        add_party_button.bind(on_press=self.add_party_popup)
        layout.add_widget(add_party_button)

        remove_party_button = Button(text="Remove Party", size_hint=(1, 0.1))
        remove_party_button.bind(on_press=self.remove_party_popup)
        layout.add_widget(remove_party_button)

        # Action selection
        self.action_var = "IN"
        action_layout = BoxLayout(orientation="horizontal", size_hint=(1, 0.1))
        in_button = Button(text="IN", on_press=lambda x: self.set_action("IN"))
        out_button = Button(text="OUT", on_press=lambda x: self.set_action("OUT"))
        action_layout.add_widget(in_button)
        action_layout.add_widget(out_button)
        layout.add_widget(action_layout)

        # Date selection
        self.date_var = datetime.now().strftime("%Y-%m-%d")
        date_layout = BoxLayout(orientation="horizontal", size_hint=(1, 0.1))
        current_date_button = Button(text="Current Date", on_press=lambda x: self.use_current_date())
        custom_date_button = Button(text="Custom Date", on_press=lambda x: self.use_custom_date())
        self.custom_date_input = TextInput(hint_text="Enter custom date (YYYY-MM-DD)", size_hint=(1, 0.1))
        date_layout.add_widget(current_date_button)
        date_layout.add_widget(custom_date_button)
        layout.add_widget(date_layout)
        layout.add_widget(self.custom_date_input)

        # Weight input
        self.weight_input = TextInput(hint_text="Enter Weight", size_hint=(1, 0.1))
        layout.add_widget(self.weight_input)

        # Car number input
        self.car_input = TextInput(hint_text="Enter Car No.", size_hint=(1, 0.1))
        layout.add_widget(self.car_input)

        # Towards party input
        self.towards_check = CheckBox(size_hint=(1, 0.1))
        self.towards_input = TextInput(hint_text="Enter Towards Party", size_hint=(1, 0.1))
        layout.add_widget(Label(text="Towards Party:"))
        layout.add_widget(self.towards_check)
        layout.add_widget(self.towards_input)

        # Description input
        self.description_check = CheckBox(size_hint=(1, 0.1))
        self.description_input = TextInput(hint_text="Enter Description", size_hint=(1, 0.1))
        layout.add_widget(Label(text="Description:"))
        layout.add_widget(self.description_check)
        layout.add_widget(self.description_input)

        # Submit button
        submit_button = Button(text="Submit", size_hint=(1, 0.2))
        submit_button.bind(on_press=self.submit)
        layout.add_widget(submit_button)

        # Back button
        back_button = Button(text="Back", size_hint=(1, 0.1))
        back_button.bind(on_press=lambda x: self.main_screen())
        layout.add_widget(back_button)

        self.root.clear_widgets()
        self.root.add_widget(layout)

    def open_ledger_screen(self, instance):
        self.clear_screen()

        layout = BoxLayout(orientation="vertical", padding=10, spacing=10)

        title = Label(text="AS Steel Inventory Ledger", font_size=20, bold=True)
        layout.add_widget(title)

        if not self.parties:
            self.show_popup("No Parties", "No parties available. Please add parties through the data entry screen.")
            self.main_screen()
            return

        # Ledger options
        ledger_layout = GridLayout(cols=2, spacing=10, size_hint=(1, 0.6))
        in_button = Button(text="IN")
        in_button.bind(on_press=lambda x: self.view_ledger("IN"))
        ledger_layout.add_widget(in_button)

        out_button = Button(text="OUT")
        out_button.bind(on_press=lambda x: self.view_ledger("OUT"))
        ledger_layout.add_widget(out_button)

        for party in self.parties:
            party_button = Button(text=party)
            party_button.bind(on_press=lambda x, p=party: self.view_ledger(p))
            ledger_layout.add_widget(party_button)

        layout.add_widget(ledger_layout)

        # Back button
        back_button = Button(text="Back", size_hint=(1, 0.1))
        back_button.bind(on_press=lambda x: self.main_screen())
        layout.add_widget(back_button)

        self.root.clear_widgets()
        self.root.add_widget(layout)

    def clear_screen(self):
        self.root.clear_widgets()

    def set_action(self, action):
        self.action_var = action

    def use_current_date(self):
        self.date_var = datetime.now().strftime("%Y-%m-%d")
        self.custom_date_input.text = ""

    def use_custom_date(self):
        self.date_var = self.custom_date_input.text

    def add_party_popup(self, instance):
        content = BoxLayout(orientation="vertical", spacing=10)
        popup_input = TextInput(hint_text="Enter new party name")
        submit_button = Button(text="Add")
        submit_button.bind(on_press=lambda x: self.add_party(popup_input.text))
        content.add_widget(popup_input)
        content.add_widget(submit_button)
        popup = Popup(title="Add Party", content=content, size_hint=(0.8, 0.4))
        popup.open()

    def add_party(self, party):
        if party and party not in self.parties:
            self.parties.append(party)
            self.party_spinner.values = self.parties
            self.save_parties()

    def remove_party_popup(self, instance):
        content = BoxLayout(orientation="vertical", spacing=10)
        popup_input = TextInput(hint_text="Enter party name to remove")
        submit_button = Button(text="Remove")
        submit_button.bind(on_press=lambda x: self.remove_party(popup_input.text))
        content.add_widget(popup_input)
        content.add_widget(submit_button)
        popup = Popup(title="Remove Party", content=content, size_hint=(0.8, 0.4))
        popup.open()

    def remove_party(self, party):
        if party in self.parties:
            self.parties.remove(party)
            self.party_spinner.values = self.parties
            self.save_parties()

    def submit(self, instance):
        party = self.party_spinner.text
        action = self.action_var
        date = self.date_var
        weight = self.weight_input.text
        car_no = self.car_input.text
        towards_party = self.towards_input.text if self.towards_check.active else "N/A"
        description = self.description_input.text if self.description_check.active else "N/A"

        if not party or not action or not weight or not car_no:
            self.show_popup("Missing Information", "Please fill out all required fields.")
            return

        filename = f"{action}_ledger.xlsx"
        if not os.path.exists(filename):
            wb = Workbook()
            ws = wb.active
            ws.append(["Party", "Date", "Weight", "Car No.", "Towards Party", "Description"])
            wb.save(filename)

        wb = load_workbook(filename)
        ws = wb.active
        ws.append([party, date, weight, car_no, towards_party, description])
        wb.save(filename)

        party_filename = f"{party}_ledger.xlsx"
        if not os.path.exists(party_filename):
            wb = Workbook()
            ws = wb.active
            ws.append(["Action", "Date", "Weight", "Car No.", "Towards Party", "Description"])
            wb.save(party_filename)

        wb = load_workbook(party_filename)
        ws = wb.active
        ws.append([action, date, weight, car_no, towards_party, description])
        wb.save(party_filename)

        self.show_popup("Success", "Data entered successfully.")

    def view_ledger(self, ledger_type):
        filename = f"{ledger_type}_ledger.xlsx" if ledger_type in ["IN", "OUT"] else f"{ledger_type}_ledger.xlsx"
        if not os.path.exists(filename):
            self.show_popup("File Not Found", f"No data found for {ledger_type}.")
            return

        wb = load_workbook(filename)
        ws = wb.active

        layout = BoxLayout(orientation="vertical")
        scroll_view = ScrollView()
        grid = GridLayout(cols=ws.max_column, size_hint_y=None)
        grid.bind(minimum_height=grid.setter('height'))

        for row in ws.iter_rows(values_only=True):
            for cell in row:
                grid.add_widget(Label(text=str(cell)))

        scroll_view.add_widget(grid)
        layout.add_widget(scroll_view)

        back_button = Button(text="Back", size_hint=(1, 0.1))
        back_button.bind(on_press=lambda x: self.open_ledger_screen(None))
        layout.add_widget(back_button)

        self.root.clear_widgets()
        self.root.add_widget(layout)

    def show_popup(self, title, message):
        popup = Popup(title=title, content=Label(text=message), size_hint=(0.8, 0.4))
        popup.open()

if __name__ == "__main__":
    SteelInventoryApp().run()